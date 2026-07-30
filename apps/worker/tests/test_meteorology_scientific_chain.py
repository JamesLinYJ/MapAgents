# +-------------------------------------------------------------------------
#
#   地理智能平台 - 气象科学交付链测试
#
#   文件:       test_meteorology_scientific_chain.py
#
#   日期:       2026年07月18日
#   作者:       JamesLinYJ
#   协助:       OpenAI Codex:GPT-5.6 Sol
# --------------------------------------------------------------------------

from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any
from zipfile import ZipFile

REPO_ROOT = Path(__file__).resolve().parents[3]
WORKER_SRC = REPO_ROOT / "apps" / "worker" / "src"
GIS_SRC = REPO_ROOT / "packages" / "gis-meteorology" / "src"
for source in (WORKER_SRC, GIS_SRC):
    if str(source) not in sys.path:
        sys.path.insert(0, str(source))

from worker_app.path_sandbox import WorkerPathSandbox
from worker_app.tool_context import WorkerToolContext
from worker_app.tool_registry import WorkerToolRegistry
from worker_app.tools import register_builtin_tools


class MeteorologyScientificChainTests(unittest.TestCase):
    def test_generates_verifiable_map_table_report_data_and_conclusion(self) -> None:
        registry = WorkerToolRegistry()
        register_builtin_tools(registry)

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            inputs = root / "inputs"
            inputs.mkdir()
            first = inputs / "202607181500_202607181505_QPF.nc"
            second = inputs / "202607181505_202607181510_QPF.nc"
            boundary = inputs / "verification_regions.geojson"
            _write_dataset(first, increment=0.0)
            _write_dataset(second, increment=2.0)
            _write_boundary(boundary)
            context = WorkerToolContext(WorkerPathSandbox(root))

            risk = registry.dispatch(
                "render_rainfall_risk_map",
                {
                    "file_relative_path": _relative(root, first),
                    "file_name": first.name,
                    "variable": "QPF",
                    "boundary_relative_path": _relative(root, boundary),
                    "thresholds": [
                        {"label": "低风险", "min": 0, "max": 5, "color": "#a6d96a"},
                        {"label": "中风险", "min": 5, "max": 10, "color": "#fdae61"},
                        {"label": "高风险", "min": 10, "max": 100, "color": "#d73027"},
                    ],
                    "map_mode": "regional",
                    "aggregation": "mean",
                    "label_field": "name",
                    "title": "科学交付链降水风险区划图",
                    "output_relative_path": "artifacts/risk-map.png",
                    "output_geojson_relative_path": "artifacts/risk-regions.geojson",
                },
                context,
            )
            risk_png = root / "artifacts" / "risk-map.png"
            risk_geojson = root / "artifacts" / "risk-regions.geojson"
            self.assertGreater(risk_png.stat().st_size, 10_000)
            self.assertEqual(risk["regionSummary"]["topRegions"][0]["name"], "东区")
            layer = json.loads(risk_geojson.read_text(encoding="utf-8"))
            self.assertEqual(len(layer["features"]), 2)
            self.assertTrue(all(feature["properties"]["risk_level"] for feature in layer["features"]))

            table = registry.dispatch(
                "generate_area_rainfall_table",
                {
                    "files": [
                        {"relativePath": _relative(root, first), "name": first.name},
                        {"relativePath": _relative(root, second), "name": second.name},
                    ],
                    "boundary_relative_path": _relative(root, boundary),
                    "top_n": 2,
                    "label_field": "name",
                    "output_xlsx_relative_path": "artifacts/area-rainfall.xlsx",
                    "output_png_relative_path": "artifacts/area-rainfall.png",
                },
                context,
            )
            table_xlsx = root / "artifacts" / "area-rainfall.xlsx"
            table_png = root / "artifacts" / "area-rainfall.png"
            self.assertGreater(table_xlsx.stat().st_size, 5_000)
            self.assertGreater(table_png.stat().st_size, 5_000)
            self.assertEqual(table["regionCount"], 2)
            self.assertEqual(table["topRows"][0]["region"], "东区")
            workbook = _load_workbook(table_xlsx)
            sheet = workbook["区域累计面雨量排行表"]
            self.assertEqual(sheet["A1"].value, "短时临近降水预报——区县区域累计面雨量排行表")
            self.assertEqual({sheet["B5"].value, sheet["B6"].value}, {"东区", "西区"})
            workbook.close()

            report = registry.dispatch(
                "meteorological_report",
                {
                    "file_relative_path": _relative(root, first),
                    "file_name": first.name,
                    "interpretation_text": "东区降水强于西区；结论仅依据本次合成网格和区划统计。",
                    "output_relative_path": "artifacts/meteorology-report.docx",
                },
                context,
            )
            report_docx = root / "artifacts" / "meteorology-report.docx"
            self.assertGreater(report_docx.stat().st_size, 10_000)
            self.assertEqual(report["variableCount"], 1)
            self.assertEqual(report["statsRowCount"], 1)
            with ZipFile(report_docx) as archive:
                document = archive.read("word/document.xml").decode("utf-8")
            self.assertIn("东区降水强于西区", document)
            self.assertIn("QPF", document)

            conclusion = {
                "highestRiskRegion": risk["regionSummary"]["topRegions"][0]["name"],
                "highestAccumulatedRainfallRegion": table["topRows"][0]["region"],
                "sourceFiles": 2,
            }
            self.assertEqual(conclusion, {
                "highestRiskRegion": "东区",
                "highestAccumulatedRainfallRegion": "东区",
                "sourceFiles": 2,
            })


def _write_dataset(path: Path, *, increment: float) -> None:
    np = _np()
    xr = _xr()
    values = np.asarray([
        [1, 2, 4, 9, 12],
        [2, 3, 5, 10, 13],
        [3, 4, 6, 11, 14],
        [4, 5, 7, 12, 15],
        [5, 6, 8, 13, 16],
    ], dtype=float) + increment
    dataset = xr.Dataset(
        data_vars={
            "QPF": (
                ("lat", "lon"),
                values,
                {"units": "mm/h", "long_name": "短时降水"},
            ),
        },
        coords={
            "lat": np.asarray([30.0, 30.25, 30.5, 30.75, 31.0]),
            "lon": np.asarray([120.0, 120.25, 120.5, 120.75, 121.0]),
        },
    )
    dataset.to_netcdf(path)
    dataset.close()


def _write_boundary(path: Path) -> None:
    feature_collection = {
        "type": "FeatureCollection",
        "features": [
            _polygon_feature("西区", 119.9, 120.5),
            _polygon_feature("东区", 120.5, 121.1),
        ],
    }
    path.write_text(json.dumps(feature_collection, ensure_ascii=False), encoding="utf-8")


def _polygon_feature(name: str, west: float, east: float) -> dict[str, object]:
    return {
        "type": "Feature",
        "properties": {"name": name},
        "geometry": {
            "type": "Polygon",
            "coordinates": [[
                [west, 29.9],
                [east, 29.9],
                [east, 31.1],
                [west, 31.1],
                [west, 29.9],
            ]],
        },
    }


def _relative(root: Path, path: Path) -> str:
    return path.relative_to(root).as_posix()


def _load_workbook(path: Path) -> Any:
    from openpyxl import load_workbook

    return load_workbook(path, read_only=True, data_only=True)


def _np() -> Any:
    import numpy as np

    return np


def _xr() -> Any:
    import xarray as xr

    return xr


if __name__ == "__main__":
    unittest.main()
